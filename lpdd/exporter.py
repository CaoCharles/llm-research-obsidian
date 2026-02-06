"""
論文匯出為 Excel 模組
"""
import re
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models import Paper


def _match_keywords(paper: Paper, keywords: dict[str, int]) -> list[str]:
    """找出論文匹配到的關鍵字列表"""
    text = f"{paper.title} {paper.abstract}".lower()
    matched = []
    for keyword in keywords:
        pattern = r"\b" + re.escape(keyword.lower()) + r"\b"
        if re.search(pattern, text):
            matched.append(keyword)
    return matched


def export_to_excel(
    papers: list[Paper],
    output_path: str | Path,
    keywords: dict[str, int] | None = None,
    title: str = "LLM 評測論文清單",
) -> Path:
    """
    將論文清單匯出為 Excel 檔案

    Args:
        papers: 論文列表
        output_path: 輸出檔案路徑
        keywords: 關鍵字權重 dict（用於標記匹配的關鍵字）
        title: Excel 標題

    Returns:
        輸出檔案路徑
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Papers"

    # 定義欄位
    columns = [
        ("No.", 6),
        ("arXiv ID", 16),
        ("論文標題", 60),
        ("作者", 35),
        ("發表日期", 14),
        ("分類", 18),
        ("匹配關鍵字", 30),
        ("匹配分數", 10),
        ("摘要 (Abstract)", 80),
        ("arXiv 連結", 35),
        ("PDF 連結", 35),
    ]

    # 樣式
    header_font = Font(name="Microsoft JhengHei", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_font = Font(name="Microsoft JhengHei", size=10)
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    link_font = Font(name="Microsoft JhengHei", size=10, color="0563C1", underline="single")

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # 標題列
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(name="Microsoft JhengHei", bold=True, size=14, color="2F5496")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # 統計列
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
    stats_cell = ws.cell(
        row=2, column=1,
        value=f"共 {len(papers)} 篇論文 | 匯出時間：{datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )
    stats_cell.font = Font(name="Microsoft JhengHei", size=10, color="666666")
    stats_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    # 表頭
    header_row = 3
    for col_idx, (col_name, col_width) in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[header_row].height = 28

    # 資料列
    even_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

    for idx, paper in enumerate(papers, 1):
        row = header_row + idx

        matched_kw = _match_keywords(paper, keywords) if keywords else []

        values = [
            idx,
            paper.arxiv_id,
            paper.title,
            ", ".join(paper.authors[:5]) + ("..." if len(paper.authors) > 5 else ""),
            paper.published.strftime("%Y-%m-%d"),
            ", ".join(paper.categories),
            ", ".join(matched_kw),
            paper.score,
            paper.abstract[:500] + ("..." if len(paper.abstract) > 500 else ""),
            paper.arxiv_url,
            paper.pdf_url,
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = thin_border

            if idx % 2 == 0:
                cell.fill = even_fill

        # 超連結
        arxiv_cell = ws.cell(row=row, column=10)
        arxiv_cell.hyperlink = paper.arxiv_url
        arxiv_cell.font = link_font

        pdf_cell = ws.cell(row=row, column=11)
        pdf_cell.hyperlink = paper.pdf_url
        pdf_cell.font = link_font

        ws.row_dimensions[row].height = 60

    # 自動篩選
    if papers:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(columns))}{header_row + len(papers)}"

    # 凍結窗格（凍結表頭）
    ws.freeze_panes = f"A{header_row + 1}"

    wb.save(str(output_path))
    return output_path


def parse_date_range(date_spec: str) -> tuple[str, str]:
    """
    解析日期範圍規格

    支援格式：
    - "2026-02-05" → 單日
    - "2026-02-01:2026-02-05" → 日期範圍
    - "7d" / "7days" → 最近 7 天
    - "1m" / "1month" → 最近 1 個月
    - "3m" / "3months" → 最近 3 個月
    - "6m" / "6months" → 最近 6 個月
    - "1y" / "1year" → 最近 1 年
    - "2026" → 整年

    Returns:
        (start_date, end_date) 格式 YYYY-MM-DD
    """
    today = datetime.now()
    spec = date_spec.strip().lower()

    # 日期範圍 "YYYY-MM-DD:YYYY-MM-DD"
    if ":" in spec:
        parts = spec.split(":")
        return parts[0].strip(), parts[1].strip()

    # 相對天數 "7d" / "7days"
    m = re.match(r"^(\d+)\s*d(?:ays?)?$", spec)
    if m:
        days = int(m.group(1))
        start = today - timedelta(days=days)
        return start.strftime("%Y-%m-%d"), today.strftime("%Y-%m-%d")

    # 相對月份 "1m" / "3months"
    m = re.match(r"^(\d+)\s*m(?:onths?)?$", spec)
    if m:
        months = int(m.group(1))
        start = today - timedelta(days=months * 30)
        return start.strftime("%Y-%m-%d"), today.strftime("%Y-%m-%d")

    # 相對年份 "1y" / "1year"
    m = re.match(r"^(\d+)\s*y(?:ears?)?$", spec)
    if m:
        years = int(m.group(1))
        start = today - timedelta(days=years * 365)
        return start.strftime("%Y-%m-%d"), today.strftime("%Y-%m-%d")

    # 整年 "2026"
    m = re.match(r"^(\d{4})$", spec)
    if m:
        year = m.group(1)
        return f"{year}-01-01", f"{year}-12-31"

    # 單日 "YYYY-MM-DD"
    m = re.match(r"^\d{4}-\d{2}-\d{2}$", spec)
    if m:
        return spec, spec

    raise ValueError(
        f"無法解析日期格式: '{date_spec}'\n"
        "支援格式: YYYY-MM-DD, YYYY-MM-DD:YYYY-MM-DD, 7d, 1m, 3m, 6m, 1y, 2026"
    )
